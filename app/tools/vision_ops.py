"""
Vision-backed tools — operate on uploaded image file_ids.

These wrap app.llm.vision (which calls Kimi K2.6 multimodal) and return
results in the standard tool-result shape so the agent can chain them.
"""
from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any, Optional

from app.llm import vision as v
from app.storage import (
    input_path,
    new_file_id,
    output_path,
    public_url,
    save_bytes,
)
from app.tools.formats import detect_from_path

log = logging.getLogger("aice.vision_ops")

IMAGE_EXTS = {"png", "jpg", "jpeg", "webp", "gif", "bmp"}


def _resolve_image(file_id: str) -> Path:
    p = output_path(file_id)
    if not p.exists():
        p = input_path(file_id)
    if not p.exists():
        raise FileNotFoundError(f"file_id not found: {file_id}")
    if detect_from_path(p) not in IMAGE_EXTS:
        raise ValueError(f"file_id is not an image (got {p.suffix})")
    return p


async def ocr_image(image_file_id: str, language_hint: Optional[str] = None) -> dict[str, Any]:
    p = _resolve_image(image_file_id)
    t0 = time.time()
    text = await v.ocr_image(p, language_hint=language_hint)
    return {
        "image_file_id": image_file_id,
        "text": text,
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def describe_image(image_file_id: str, question: Optional[str] = None) -> dict[str, Any]:
    p = _resolve_image(image_file_id)
    t0 = time.time()
    desc = await v.describe_image(p, question=question)
    return {
        "image_file_id": image_file_id,
        "description": desc,
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def extract_table_from_image(image_file_id: str) -> dict[str, Any]:
    p = _resolve_image(image_file_id)
    t0 = time.time()
    table = await v.extract_table_from_image(p)
    return {
        "image_file_id": image_file_id,
        "headers": table.get("headers", []),
        "rows": table.get("rows", []),
        "row_count": len(table.get("rows", [])),
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def generate_image(
    prompt: str,
    *,
    model: str = "black-forest-labs/FLUX.2-pro",
    size: str = "1024x1024",
) -> dict[str, Any]:
    """Generate an image from a text prompt. Saves to outputs/ and returns
    image_file_id ready for insert_image / pptx insertion."""
    t0 = time.time()
    img_bytes = await v.generate_image(prompt, model=model, size=size)
    fid = new_file_id("png")
    with open(output_path(fid), "wb") as f:
        f.write(img_bytes)
    return {
        "image_file_id": fid,
        "url": public_url(fid),
        "model": model,
        "size_bytes": len(img_bytes),
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def image_to_xlsx(image_file_id: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Vision-extract a table from an image, then write it as XLSX."""
    p = _resolve_image(image_file_id)
    t0 = time.time()
    table = await v.extract_table_from_image(p)
    headers = table.get("headers", []) or []
    rows = table.get("rows", []) or []

    out_id = new_file_id("xlsx")
    out_path = output_path(out_id)
    _write_xlsx(out_path, headers, rows, sheet_name)
    return {
        "image_file_id": image_file_id,
        "file_id": out_id,
        "url": public_url(out_id),
        "headers": headers,
        "row_count": len(rows),
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


def _write_xlsx(out_path: Path, headers: list[str], rows: list[list[str]], sheet_name: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"

    if headers:
        ws.append(headers)
        bold = Font(bold=True)
        fill = PatternFill(start_color="EBDCB1", end_color="EBDCB1", fill_type="solid")
        for cell in ws[1]:
            cell.font = bold
            cell.fill = fill

    for row in rows:
        # Coerce non-list values defensively
        if not isinstance(row, (list, tuple)):
            row = [row]
        ws.append([str(c) if c is not None else "" for c in row])

    # Auto-width-ish: set column widths based on max content length
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(8, max_len + 2), 40)

    wb.save(str(out_path))
