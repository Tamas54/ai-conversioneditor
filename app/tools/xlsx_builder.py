"""
XLSX builder — agentic spreadsheet construction with formulas, styling, and
multi-sheet support.

Designed so Kimi can reproduce a spreadsheet from a photo:
  1. vision describes the layout (headers, merged cells, formula patterns,
     number formats, conditional formatting hints)
  2. extract_table_from_image gets row data
  3. agent calls these builder tools to construct the workbook with proper
     FORMULAS (not hardcoded values), styling, banding, etc.

Anthropic's xlsx skill principle: 'Always use Excel formulas instead of
hardcoded values.' These tools make that possible.
"""
from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any, Optional, Union

from app.storage import (
    input_path,
    new_file_id,
    output_path,
    public_url,
)

log = logging.getLogger("aice.xlsx_builder")


def _resolve(file_id: str) -> Path:
    p = output_path(file_id)
    if p.exists():
        return p
    p = input_path(file_id)
    if p.exists():
        return p
    raise FileNotFoundError(f"file_id not found: {file_id}")


def _new_xlsx_output() -> tuple[str, Path]:
    fid = new_file_id("xlsx")
    return fid, output_path(fid)


def _hex(color: str) -> str:
    """Normalize a hex color to 'RRGGBB' (no #) for openpyxl."""
    return color.strip().lstrip("#").upper()


def _coerce_value(v: Any) -> Any:
    """Coerce a string-shaped number/percentage back to a numeric type for
    Excel-friendly storage. '12,5%' → 0.125, '1 234' → 1234."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return v
    s = v.strip()
    if not s:
        return None
    # Don't coerce if it looks like a formula
    if s.startswith("="):
        return s
    # Percentage
    pct = False
    if s.endswith("%"):
        pct = True
        s = s[:-1].strip()
    s_clean = s.replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        f = float(s_clean)
        if pct:
            return f / 100
        return int(f) if f.is_integer() else f
    except ValueError:
        return v


# ============================================================
# Builder primitives
# ============================================================


async def xlsx_create(sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create a new empty workbook with one sheet."""
    from openpyxl import Workbook

    t0 = time.time()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "active_sheet": ws.title,
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_add_sheet(file_id: str, name: str, *, position: Optional[int] = None) -> dict[str, Any]:
    from openpyxl import load_workbook
    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    ws = wb.create_sheet(title=name[:31], index=position)
    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "added_sheet": ws.title,
        "sheets": wb.sheetnames,
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_write_block(
    file_id: str,
    sheet: str,
    data: list[list[Any]],
    *,
    anchor: str = "A1",
    headers: Optional[list[str]] = None,
    coerce_numeric: bool = True,
) -> dict[str, Any]:
    """Bulk-write a 2D array starting at `anchor`. If `headers` is provided,
    they're written as the first row at anchor, followed by data rows below.
    Numeric strings ('12,5%', '1 234') are coerced unless coerce_numeric=False."""
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string

    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet!r}")
    ws = wb[sheet]
    col_letter, start_row = coordinate_from_string(anchor)
    start_col = column_index_from_string(col_letter)

    cur_row = start_row
    if headers:
        for j, h in enumerate(headers):
            ws.cell(row=cur_row, column=start_col + j, value=str(h))
        cur_row += 1

    n_data = 0
    for row in data:
        for j, v in enumerate(row):
            val = _coerce_value(v) if coerce_numeric else v
            ws.cell(row=cur_row, column=start_col + j, value=val)
        cur_row += 1
        n_data += 1

    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "rows_written": n_data + (1 if headers else 0),
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_set_cell(
    file_id: str, sheet: str, cell: str, value: Any,
    *, number_format: Optional[str] = None,
    coerce_numeric: bool = True,
) -> dict[str, Any]:
    """Set a single cell. If `value` starts with '=' it's stored as a formula.
    Pass `number_format` like '#,##0', '0.0%', '#,##0.00 [$USD]'."""
    from openpyxl import load_workbook

    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet!r}")
    ws = wb[sheet]
    val = _coerce_value(value) if coerce_numeric else value
    ws[cell] = val
    if number_format:
        ws[cell].number_format = number_format
    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "cell": cell, "stored": str(val)[:60],
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_set_formula(
    file_id: str, sheet: str, cell: str, formula: str,
    *, number_format: Optional[str] = None,
) -> dict[str, Any]:
    """Set an explicit formula in a cell. `formula` may include or omit the
    leading '=' — it's added if missing."""
    if not formula.startswith("="):
        formula = "=" + formula
    return await xlsx_set_cell(file_id, sheet, cell, formula,
                               number_format=number_format,
                               coerce_numeric=False)


async def xlsx_apply_style(
    file_id: str, sheet: str, range_ref: str,
    *,
    bold: Optional[bool] = None,
    italic: Optional[bool] = None,
    font_size: Optional[float] = None,
    font_family: Optional[str] = None,
    font_color: Optional[str] = None,         # hex or name
    fill_color: Optional[str] = None,         # hex
    horizontal: Optional[str] = None,         # 'left'|'center'|'right'
    vertical: Optional[str] = None,           # 'top'|'center'|'bottom'
    number_format: Optional[str] = None,
    border_all: Optional[str] = None,         # hex; thin border around all cells
    wrap_text: Optional[bool] = None,
) -> dict[str, Any]:
    """Apply font / fill / alignment / number-format / border to a range
    (e.g. 'A1:D10' or single cell 'B5')."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color

    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet!r}")
    ws = wb[sheet]

    cells = ws[range_ref]
    # Normalize: ws[range] returns tuple-of-tuples for ranges, single Cell for a cell ref
    if not isinstance(cells, tuple):
        cells = ((cells,),)
    if cells and not isinstance(cells[0], tuple):
        cells = (cells,)

    side = Side(style="thin", color=_hex(border_all)) if border_all else None
    border = Border(left=side, right=side, top=side, bottom=side) if side else None

    n = 0
    for row in cells:
        for c in row:
            # Font
            if bold is not None or italic is not None or font_size is not None \
                    or font_family or font_color:
                f = c.font
                c.font = Font(
                    name=font_family or f.name,
                    size=font_size if font_size is not None else f.size,
                    bold=bold if bold is not None else f.bold,
                    italic=italic if italic is not None else f.italic,
                    color=Color(rgb="FF" + _hex(font_color)) if font_color else f.color,
                )
            # Fill
            if fill_color:
                c.fill = PatternFill(
                    start_color="FF" + _hex(fill_color),
                    end_color="FF" + _hex(fill_color),
                    fill_type="solid",
                )
            # Alignment
            if horizontal or vertical or wrap_text is not None:
                a = c.alignment
                c.alignment = Alignment(
                    horizontal=horizontal or a.horizontal,
                    vertical=vertical or a.vertical,
                    wrap_text=wrap_text if wrap_text is not None else a.wrap_text,
                )
            if number_format:
                c.number_format = number_format
            if border:
                c.border = border
            n += 1

    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "cells_styled": n,
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_merge_cells(
    file_id: str, sheet: str, range_ref: str,
) -> dict[str, Any]:
    from openpyxl import load_workbook
    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    ws = wb[sheet]
    ws.merge_cells(range_ref)
    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "merged": range_ref,
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_set_column_widths(
    file_id: str, sheet: str, widths: dict[str, float],
) -> dict[str, Any]:
    """widths: {'A': 18, 'B': 12, ...}"""
    from openpyxl import load_workbook
    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    ws = wb[sheet]
    for col, w in widths.items():
        ws.column_dimensions[col.upper()].width = float(w)
    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {
        "file_id": fid, "url": public_url(fid),
        "columns_set": list(widths.keys()),
        "ms_elapsed": int((time.time() - t0) * 1000),
    }


async def xlsx_freeze_panes(
    file_id: str, sheet: str, anchor: str = "B2",
) -> dict[str, Any]:
    """Freeze rows above and columns left of `anchor` (e.g. 'B2' freezes
    row 1 and column A)."""
    from openpyxl import load_workbook
    src = _resolve(file_id)
    t0 = time.time()
    wb = load_workbook(str(src))
    ws = wb[sheet]
    ws.freeze_panes = anchor
    fid, dst = _new_xlsx_output()
    wb.save(str(dst))
    return {"file_id": fid, "url": public_url(fid), "frozen_at": anchor,
            "ms_elapsed": int((time.time() - t0) * 1000)}


async def xlsx_inspect(
    file_id: str, sheet: Optional[str] = None,
    max_rows: int = 25, max_cols: int = 12,
) -> dict[str, Any]:
    """Read back what's in the workbook — for verification after building."""
    from openpyxl import load_workbook
    src = _resolve(file_id)
    wb = load_workbook(str(src), data_only=False)
    if sheet is None:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=False), start=1):
        row_data = []
        for c in row:
            if c.value is None:
                row_data.append("")
            elif isinstance(c.value, str) and c.value.startswith("="):
                row_data.append(c.value)  # formula
            else:
                row_data.append(c.value)
        rows.append(row_data)
    return {
        "file_id": file_id,
        "sheets": wb.sheetnames,
        "active_sheet": sheet,
        "shown_rows": rows,
        "merged_ranges": [str(r) for r in ws.merged_cells.ranges],
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }
